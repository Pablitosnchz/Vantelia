# -*- coding: utf-8 -*-
"""Importa el catalogo de un salon desde su Excel de servicios y packs.

Nace del alta de un cliente real (ago 2026) que entrego su tabla de gestion con
164 servicios y 28 packs. Teclearlo a mano no es viable, y copiarlo mal es peor:
la duracion es lo que bloquea la agenda y el precio es lo que se cobra.

Formato esperado, hoja "Servicios":
    Codigo | Servicio | Categoria | Precios | Tiempo de realizacion (min) | Operario | Descripcion

Formato esperado, hoja "Packs":
    Nombre del pack | Paso 1 | Tiempo exposicion 1 | Paso 2 | Tiempo 2 | ... (hasta 10)

Reglas que aplica, y por que:

* Un PACK es UNA cita, no varias: se crea como un servicio mas, con la duracion
  sumada de sus pasos (el tiempo de exposicion si lo indican, si no la duracion
  propia de ese servicio) y el precio sumado de sus componentes.
* Las filas "FIANZA ..." NO son servicios que se reserven: dicen cuanto pide el
  salon por adelantado y para que trabajos. Se traducen a la senal de los
  servicios afectados (payment_type='deposit').
* Los pasos internos que no se venden sueltos (valen 0 EUR y solo aparecen
  dentro de packs) quedan INACTIVOS: siguen existiendo para que la duracion del
  pack cuadre, pero no se le ofrecen al cliente.
* "Operario" dice quien hace cada servicio. "Todos" = sin restriccion.

Uso:
    python scripts/importar_catalogo_excel.py <excel> <cliente_id> [--aplicar]

Sin --aplicar solo enseña lo que haria. Conviene mirarlo antes de tocar nada.
"""
from __future__ import annotations

import argparse
import re
import sys
import time
import unicodedata
from typing import Any, Dict, List, Tuple

FIANZA_PREFIJO = "FIANZA"


def _slug(nombre: str) -> str:
    base = unicodedata.normalize("NFKD", nombre).encode("ascii", "ignore").decode()
    base = re.sub(r"[^a-zA-Z0-9]+", "_", base).strip("_").lower()
    return base[:80] or "servicio"


def _titulo(nombre: str) -> str:
    """MAYUSCULAS DE GESTION -> texto legible para el cliente final."""
    limpio = re.sub(r"\s+", " ", str(nombre or "").strip())
    return limpio[:1].upper() + limpio[1:].lower() if limpio else ""


def _euros_a_centimos(valor: Any) -> int:
    try:
        return int(round(float(valor) * 100))
    except (TypeError, ValueError):
        return 0


def _minutos(valor: Any, defecto: int = 30) -> int:
    try:
        n = int(round(float(valor)))
    except (TypeError, ValueError):
        return defecto
    return n if n > 0 else defecto


def leer_excel(ruta: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Devuelve (servicios, packs, fianzas) tal y como vienen en el Excel."""
    # Import perezoso: `aplicar` se usa desde el servidor, donde no hace falta leer
    # ningun Excel y no tiene sentido exigir la dependencia.
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - dependencia de uso puntual
        sys.exit("Falta openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(ruta, data_only=True)
    servicios: List[Dict[str, Any]] = []
    fianzas: List[Dict[str, Any]] = []
    for fila in wb["Servicios"].iter_rows(min_row=2, values_only=True):
        nombre = str(fila[1] or "").strip()
        if not nombre:
            continue
        registro = {
            "nombre": nombre,
            "categoria": str(fila[2] or "").strip(),
            "precio_cents": _euros_a_centimos(fila[3]),
            "minutos": _minutos(fila[4]),
            "operario": str(fila[5] or "Todos").strip(),
            "descripcion": str(fila[6] or "").strip(),
        }
        destino = fianzas if nombre.upper().startswith(FIANZA_PREFIJO) else servicios
        destino.append(registro)

    packs: List[Dict[str, Any]] = []
    for fila in wb["Packs"].iter_rows(min_row=2, values_only=True):
        nombre = str(fila[0] or "").strip()
        if not nombre:
            continue
        pasos = []
        for i in range(1, 21, 2):
            paso = str(fila[i] or "").strip() if i < len(fila) else ""
            if paso:
                pasos.append((paso, fila[i + 1] if i + 1 < len(fila) else None))
        if pasos:
            packs.append({"nombre": nombre, "pasos": pasos})
    return servicios, packs, fianzas


def componer_packs(
    packs: List[Dict[str, Any]], servicios: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Un pack = un servicio con la duracion y el precio de sus pasos sumados."""
    por_nombre = {s["nombre"].upper(): s for s in servicios}
    compuestos: List[Dict[str, Any]] = []
    sin_resolver: List[str] = []
    usados_en_packs = set()
    for pack in packs:
        minutos = 0
        precio = 0
        detalle: List[str] = []
        for paso, exposicion in pack["pasos"]:
            base = por_nombre.get(paso.upper())
            if base is None:
                sin_resolver.append("%s -> %s" % (pack["nombre"], paso))
                continue
            usados_en_packs.add(base["nombre"].upper())
            # Celda VACIA = no lo indican, se usa la duracion propia del servicio.
            # Un CERO es un dato: significa que ese paso no tiene espera. Confundirlos
            # inflaba un pack a 10 horas, porque el paso valia 0 y se le sumaban los
            # 300 minutos del servicio completo.
            if exposicion is None or str(exposicion).strip() == "":
                minutos += base["minutos"]
            else:
                minutos += max(0, _minutos(exposicion, 0))
            precio += base["precio_cents"]
            detalle.append(_titulo(base["nombre"]))
        # Hay packs cuyos pasos valen 0 EUR porque el precio real vive en otro
        # servicio del catalogo (los de keratina y acido lactico). Sumar da cifras
        # absurdas --un tratamiento de dos horas por 4 EUR-- asi que se deja "a
        # consultar" antes que publicar un precio inventado.
        if precio < minutos * 30:  # menos de 0,30 EUR por minuto no se sostiene
            precio = 0
        compuestos.append({
            "nombre": pack["nombre"],
            "categoria": "PACKS",
            "precio_cents": precio,
            "minutos": minutos or 60,
            "operario": "Todos",
            "descripcion": ("Incluye: " + " + ".join(detalle)) if detalle else "",
        })
    for servicio in servicios:
        servicio["solo_en_pack"] = (
            servicio["precio_cents"] == 0 and servicio["nombre"].upper() in usados_en_packs
        )
    return compuestos, sin_resolver


def senal_para(servicio: Dict[str, Any], fianzas: List[Dict[str, Any]]) -> int:
    """Cuanta senal pide este servicio, segun las filas FIANZA del propio Excel."""
    # Solo el nombre y la categoria. La descripcion mete falsos positivos: "Elumen"
    # se describe como "coloracion PERMANENTE sin amoniaco" y acababa pidiendo senal
    # por la palabra permanente, que ahi significa otra cosa.
    texto = (servicio["nombre"] + " " + servicio["categoria"]).upper()
    mejor = 0
    for fianza in fianzas:
        etiquetas = [
            t.strip()
            for t in re.split(r"[-,/]| Y ", fianza["nombre"].upper().replace(FIANZA_PREFIJO, "", 1))
            if len(t.strip()) > 3
        ]
        for etiqueta in etiquetas:
            # DECOLORACIONES casa con DECOLORACION, EXTENSIONES con EXTENSION.
            raiz = etiqueta[:-2] if etiqueta.endswith("ES") else etiqueta
            if raiz and raiz in texto:
                mejor = max(mejor, fianza["precio_cents"])
                break
    return mejor


def equipo_del_excel(servicios: List[Dict[str, Any]]) -> List[str]:
    """Nombres reales del equipo, sacados de la columna Operario."""
    nombres: List[str] = []
    for servicio in servicios:
        for parte in servicio["operario"].split(","):
            nombre = parte.strip()
            if nombre and nombre.lower() != "todos" and nombre not in nombres:
                nombres.append(nombre)
    return nombres


def preparar(ruta: str) -> Dict[str, Any]:
    servicios, packs, fianzas = leer_excel(ruta)
    compuestos, sin_resolver = componer_packs(packs, servicios)
    catalogo = servicios + compuestos
    for servicio in catalogo:
        servicio["senal_cents"] = senal_para(servicio, fianzas)
        servicio["slug"] = _slug(servicio["nombre"])
    return {
        "catalogo": catalogo,
        "fianzas": fianzas,
        "equipo": equipo_del_excel(servicios),
        "sin_resolver": sin_resolver,
    }


def pide_senal(servicio: Dict[str, Any]) -> bool:
    """Solo tiene sentido pedir senal si el servicio cuesta mas que la propia senal."""
    senal = int(servicio.get("senal_cents") or 0)
    return bool(senal and servicio["precio_cents"] > senal)


def aplicar(datos: Dict[str, Any], cliente_id: str, connection) -> Dict[str, int]:
    """Deja el catalogo del tenant EXACTAMENTE como el Excel. Idempotente."""
    ahora = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    connection.execute("DELETE FROM services WHERE cliente_id=?", (cliente_id,))
    orden_categoria: Dict[str, int] = {}
    for servicio in datos["catalogo"]:
        orden_categoria.setdefault(servicio["categoria"], len(orden_categoria))
    vistos = set()
    creados = 0
    for i, servicio in enumerate(datos["catalogo"]):
        slug = servicio["slug"]
        if slug in vistos:  # nombres repetidos en el Excel
            continue
        vistos.add(slug)
        con_senal = pide_senal(servicio)
        connection.execute(
            """
            INSERT INTO services (cliente_id, slug, name, duration_minutes, price_cents,
                description, is_active, sort_order, payment_mode, payment_type,
                deposit_amount_cents, currency, category, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'eur', ?, ?, ?)
            """,
            (
                cliente_id, slug, _titulo(servicio["nombre"]),
                servicio["minutos"], servicio["precio_cents"], servicio["descripcion"],
                0 if servicio.get("solo_en_pack") else 1,
                orden_categoria.get(servicio["categoria"], 99) * 1000 + i,
                "payment_required" if con_senal else "payment_disabled",
                "deposit" if con_senal else "full",
                servicio["senal_cents"] if con_senal else 0,
                _titulo(servicio["categoria"]), ahora, ahora,
            ),
        )
        creados += 1
    return {"servicios": creados}


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa servicios y packs desde Excel.")
    parser.add_argument("excel")
    parser.add_argument("cliente_id")
    parser.add_argument("--aplicar", action="store_true", help="escribe en la base de datos")
    parser.add_argument("--db", default="storage/vantelia.db")
    args = parser.parse_args()

    datos = preparar(args.excel)
    catalogo = datos["catalogo"]
    activos = [s for s in catalogo if not s.get("solo_en_pack")]
    con_senal = [s for s in activos if pide_senal(s)]
    print("Servicios en el Excel  : %d" % len(catalogo))
    print("  se ofrecen al cliente: %d" % len(activos))
    print("  pasos internos (off) : %d" % (len(catalogo) - len(activos)))
    print("  con senal            : %d" % len(con_senal))
    print("Equipo detectado       : %s" % ", ".join(datos["equipo"]))
    print("Fianzas del Excel      : %s" % "; ".join(
        "%s = %.2f EUR" % (f["nombre"].strip(), f["precio_cents"] / 100) for f in datos["fianzas"]))
    if con_senal:
        print("\nEjemplos con senal:")
        for servicio in con_senal[:8]:
            print("   - %-52s %6.2f EUR  senal %.2f" % (
                _titulo(servicio["nombre"])[:52], servicio["precio_cents"] / 100,
                servicio["senal_cents"] / 100))
    if datos["sin_resolver"]:
        print("\nPasos de pack que no casan con ningun servicio (%d):" % len(datos["sin_resolver"]))
        for aviso in datos["sin_resolver"]:
            print("   - %s" % aviso)

    if not args.aplicar:
        print("\n(simulacion: nada escrito; usa --aplicar para guardar)")
        return

    import sqlite3

    connection = sqlite3.connect(args.db)
    try:
        resultado = aplicar(datos, args.cliente_id, connection)
        connection.commit()
    finally:
        connection.close()
    print("\nEscrito en %s: %d servicios para %s" % (args.db, resultado["servicios"], args.cliente_id))


if __name__ == "__main__":
    main()
